"""
pnl_engine.py
-------------
Pure business-logic module for PNL calculation.
No Streamlit dependency — can be unit-tested independently.

Settlement scope:
  • NFO → NIFTY only (filtered upstream by bhavcopy_provider)
  • BFO → SENSEX only (filtered upstream by bhavcopy_provider)
"""

import logging
import re
from io import BytesIO

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

logger = logging.getLogger(__name__)


# ═════════════════════════════════════════════════════════════════════
# Excel helper
# ═════════════════════════════════════════════════════════════════════
def styled_excel_bytes(df: pd.DataFrame, sheet_name: str) -> bytes:
    """Write a DataFrame to a styled Excel workbook and return as bytes."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        for row in ws.rows:
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", fill_type="solid")

    return output.getvalue()


# ═════════════════════════════════════════════════════════════════════
# Position normalisation
# ═════════════════════════════════════════════════════════════════════
_REQUIRED_POSITION_COLS = [
    "UserID", "Exchange", "Symbol", "Net Qty",
    "Buy Avg Price", "Sell Avg Price",
    "Sell Qty", "Buy Qty",
    "Realized Profit", "Unrealized Profit",
]


def validate_positions_columns(df: pd.DataFrame) -> None:
    """Raise ValueError if any required column is missing."""
    missing = [c for c in _REQUIRED_POSITION_COLS if c not in df.columns]
    if missing:
        raise ValueError(f"Missing columns in positions file: {missing}")


def normalize_positions_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Clean and normalise a raw positions DataFrame.

    • Strips column names
    • Coerces numeric fields
    • Normalises F&O symbols to '<Strike><CE/PE>' format
    """
    df = df.copy()
    df.columns = df.columns.str.strip()
    validate_positions_columns(df)

    numeric_cols = ["Net Qty", "Buy Avg Price", "Sell Avg Price", "Sell Qty", "Buy Qty"]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    df["Exchange"] = df["Exchange"].astype(str).str.strip().str.upper()
    df["Symbol"] = df["Symbol"].astype(str).str.strip()

    mask = df["Exchange"].isin(["NFO", "BFO"])
    has_space = df.loc[mask, "Symbol"].str.contains(" ", na=False)

    # With space: e.g. "NIFTY 21AUG2025 PE 24150" → "24150PE"
    df.loc[mask & has_space, "Symbol"] = (
        df.loc[mask & has_space, "Symbol"].str[-5:]
        + df.loc[mask & has_space, "Symbol"].str[-8:-6]
    )

    # Without space: last 7 chars → e.g. "NIFTY2610626850CE" → "26850CE"
    df.loc[mask & ~has_space, "Symbol"] = (
        df.loc[mask & ~has_space, "Symbol"].str[-7:]
    )

    return df


# ═════════════════════════════════════════════════════════════════════
# PNL calculation
# ═════════════════════════════════════════════════════════════════════
def enrich_positions_with_pnl(
    df: pd.DataFrame,
    nfo_bhav_df: pd.DataFrame | None,
    bfo_bhav_df: pd.DataFrame | None,
    include_settlement_nfo: bool,
    include_settlement_bfo: bool,
) -> pd.DataFrame:
    """
    Compute Realized PNL and Settlement PNL for every position row.

    Parameters
    ----------
    df                    : Raw positions DataFrame
    nfo_bhav_df           : NFO settlement data [Strike_Type, SETTLEMENT]  or None
    bfo_bhav_df           : BFO settlement data [Symbols, Close Price]     or None
    include_settlement_nfo: Whether to include NFO settlement
    include_settlement_bfo: Whether to include BFO settlement

    Returns
    -------
    Enriched DataFrame with PNL columns added.
    """
    df = normalize_positions_df(df)

    df["Calculated_Realized_PNL"] = 0.0
    df["Calculated_Settlement_PNL"] = 0.0
    df["SETTLEMENT"] = np.nan
    df["Close Price"] = np.nan
    df["Strike_Type"] = np.nan

    df_nfo_mask = df["Exchange"] == "NFO"
    df_bfo_mask = df["Exchange"] == "BFO"

    # ── Realized PNL ─────────────────────────────────────────────────
    for exchange_mask in [df_nfo_mask, df_bfo_mask]:
        sub = df.loc[exchange_mask].copy()
        conditions = [sub["Net Qty"] == 0, sub["Net Qty"] > 0, sub["Net Qty"] < 0]
        choices = [
            (sub["Sell Avg Price"] - sub["Buy Avg Price"]) * sub["Sell Qty"],
            (sub["Sell Avg Price"] - sub["Buy Avg Price"]) * sub["Sell Qty"],
            (sub["Sell Avg Price"] - sub["Buy Avg Price"]) * sub["Buy Qty"],
        ]
        df.loc[exchange_mask, "Calculated_Realized_PNL"] = np.select(
            conditions, choices, default=0
        )

    # ── NFO Settlement ───────────────────────────────────────────────
    df.loc[df_nfo_mask, "Strike_Type"] = (
        df.loc[df_nfo_mask, "Symbol"]
        .str.extract(r"(\d+[A-Z]{2})$")[0]
        .astype("string")
    )

    if include_settlement_nfo and nfo_bhav_df is not None:
        merged_nfo = df.loc[df_nfo_mask, ["Strike_Type"]].merge(
            nfo_bhav_df, on="Strike_Type", how="left"
        )
        df.loc[df_nfo_mask, "SETTLEMENT"] = merged_nfo["SETTLEMENT"].values
    elif not include_settlement_nfo:
        df.loc[df_nfo_mask, "SETTLEMENT"] = 0

    df.loc[df_nfo_mask, "SETTLEMENT"] = pd.to_numeric(
        df.loc[df_nfo_mask, "SETTLEMENT"], errors="coerce"
    ).fillna(0)

    df.loc[df_nfo_mask, "Calculated_Settlement_PNL"] = np.select(
        [df.loc[df_nfo_mask, "Net Qty"] > 0, df.loc[df_nfo_mask, "Net Qty"] < 0],
        [
            (df.loc[df_nfo_mask, "SETTLEMENT"] - df.loc[df_nfo_mask, "Buy Avg Price"])
            * df.loc[df_nfo_mask, "Net Qty"].abs(),
            (df.loc[df_nfo_mask, "Sell Avg Price"] - df.loc[df_nfo_mask, "SETTLEMENT"])
            * df.loc[df_nfo_mask, "Net Qty"].abs(),
        ],
        default=0,
    )

    # ── BFO Settlement ───────────────────────────────────────────────
    if include_settlement_bfo and bfo_bhav_df is not None:
        close_map = (
            bfo_bhav_df.drop_duplicates("Symbols")
            .set_index("Symbols")["Close Price"]
        )
        df.loc[df_bfo_mask, "Close Price"] = (
            df.loc[df_bfo_mask, "Symbol"].str.strip().map(close_map)
        )
    elif not include_settlement_bfo:
        df.loc[df_bfo_mask, "Close Price"] = 0

    df.loc[df_bfo_mask, "Close Price"] = pd.to_numeric(
        df.loc[df_bfo_mask, "Close Price"], errors="coerce"
    ).fillna(0)

    df.loc[df_bfo_mask, "Calculated_Settlement_PNL"] = np.select(
        [df.loc[df_bfo_mask, "Net Qty"] > 0, df.loc[df_bfo_mask, "Net Qty"] < 0],
        [
            (df.loc[df_bfo_mask, "Close Price"] - df.loc[df_bfo_mask, "Buy Avg Price"])
            * df.loc[df_bfo_mask, "Net Qty"].abs(),
            (df.loc[df_bfo_mask, "Sell Avg Price"] - df.loc[df_bfo_mask, "Close Price"])
            * df.loc[df_bfo_mask, "Net Qty"].abs(),
        ],
        default=0,
    )

    # ── Totals ───────────────────────────────────────────────────────
    df["Calculated_Realized_PNL"] = pd.to_numeric(
        df["Calculated_Realized_PNL"], errors="coerce"
    ).fillna(0)
    df["Calculated_Settlement_PNL"] = pd.to_numeric(
        df["Calculated_Settlement_PNL"], errors="coerce"
    ).fillna(0)
    df["Row Grand Total"] = df["Calculated_Realized_PNL"] + df["Calculated_Settlement_PNL"]

    return df


# ═════════════════════════════════════════════════════════════════════
# Summary file parser (Users sheet → Allocation + MTM)
# ═════════════════════════════════════════════════════════════════════
def parse_summary_file(file_obj) -> tuple[pd.DataFrame | None, list[str]]:
    """
    Parse the summary Excel file's **Users** sheet.

    Extracts per-user:
      • UserID         — auto-detected from first/matching column
      • Allocation×100 — from the ALLOCATION column
      • MTM            — numeric value after ``MTM=`` in the REMARK column

    Returns
    -------
    (DataFrame | None, log_lines)
        DataFrame columns: [UserID, Allocation, MTM]
    """
    logs: list[str] = []

    try:
        file_obj.seek(0)
    except Exception:
        pass

    try:
        xl = pd.ExcelFile(file_obj)
    except Exception as e:
        logs.append(f"❌ Failed to open summary file: {e}")
        return None, logs

    # ── Find the Users sheet ─────────────────────────────────────────
    users_sheet = None
    for sheet in xl.sheet_names:
        if "user" in sheet.lower():
            users_sheet = sheet
            break
    if users_sheet is None:
        users_sheet = xl.sheet_names[0]
        logs.append(f"⚠️ No 'Users' sheet found — using first sheet: {users_sheet}")

    df = xl.parse(users_sheet)
    df.columns = df.columns.str.strip()
    logs.append(f"📋 Read '{users_sheet}' sheet — {len(df)} rows, columns: {list(df.columns)}")

    # ── Auto-detect UserID column ────────────────────────────────────
    userid_col = None
    for col in df.columns:
        normalized = col.lower().replace(" ", "").replace("_", "")
        if normalized in ("userid", "user_id", "username", "user"):
            userid_col = col
            break
    if userid_col is None:
        # Fallback: first column
        userid_col = df.columns[0]
        logs.append(f"ℹ️ No explicit UserID column found — using '{userid_col}'")
    else:
        logs.append(f"✅ UserID column: {userid_col}")

    # ── Find ALLOCATION column ───────────────────────────────────────
    alloc_col = None
    for col in df.columns:
        if "allocation" in col.lower():
            alloc_col = col
            break

    # ── Find REMARK column ───────────────────────────────────────────
    remark_col = None
    for col in df.columns:
        if "remark" in col.lower():
            remark_col = col
            break

    # ── Build result ─────────────────────────────────────────────────
    result = pd.DataFrame()
    result["UserID"] = df[userid_col].astype(str).str.strip()

    if alloc_col:
        result["Allocation"] = (
            pd.to_numeric(df[alloc_col], errors="coerce").fillna(0) * 100
        )
        logs.append(f"✅ Allocation column: {alloc_col} (×100 applied)")
    else:
        result["Allocation"] = 0
        logs.append("⚠️ No ALLOCATION column found — defaulting to 0")

    if remark_col:
        mtm_raw = df[remark_col].astype(str).str.extract(
            r"MTM=([+-]?\d+\.?\d*)", expand=False
        )
        result["MTM"] = pd.to_numeric(mtm_raw, errors="coerce").fillna(0)
        logs.append(f"✅ MTM extracted from: {remark_col}")
    else:
        result["MTM"] = 0
        logs.append("⚠️ No REMARK column found — MTM defaulting to 0")

    # Drop rows where UserID is empty/nan
    result = result[result["UserID"].notna() & (result["UserID"] != "") & (result["UserID"] != "nan")]

    logs.append(f"✅ Summary parsed — {len(result)} user(s)")
    return result, logs


def build_user_summary(
    processed_df: pd.DataFrame,
    summary_data: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """
    Aggregate PNL by UserID across exchanges.

    If ``summary_data`` is provided (from parse_summary_file), the
    Allocation and MTM columns are merged into the final result.
    """
    df = processed_df.copy()
    df["NFO Realized"] = np.where(df["Exchange"] == "NFO", df["Calculated_Realized_PNL"], 0)
    df["NFO Settlement"] = np.where(df["Exchange"] == "NFO", df["Calculated_Settlement_PNL"], 0)
    df["BFO Realized"] = np.where(df["Exchange"] == "BFO", df["Calculated_Realized_PNL"], 0)
    df["BFO Settlement"] = np.where(df["Exchange"] == "BFO", df["Calculated_Settlement_PNL"], 0)

    summary_df = (
        df.groupby("UserID", dropna=False)[
            ["NFO Realized", "NFO Settlement", "BFO Realized", "BFO Settlement"]
        ]
        .sum()
        .reset_index()
    )

    summary_df["Total Realized"] = summary_df["NFO Realized"] + summary_df["BFO Realized"]
    summary_df["Total Settlement"] = summary_df["NFO Settlement"] + summary_df["BFO Settlement"]
    summary_df["Grand Total"] = summary_df["Total Realized"] + summary_df["Total Settlement"]

    # ── Merge summary file data (Allocation, MTM) ────────────────────
    if summary_data is not None and not summary_data.empty:
        summary_data = summary_data.copy()
        summary_data["UserID"] = summary_data["UserID"].astype(str).str.strip()
        summary_df["UserID"] = summary_df["UserID"].astype(str).str.strip()

        summary_df = summary_df.merge(
            summary_data[["UserID", "Allocation", "MTM"]],
            on="UserID",
            how="left",
        )
        summary_df["Allocation"] = summary_df["Allocation"].fillna(0)
        summary_df["MTM"] = summary_df["MTM"].fillna(0)
    else:
        summary_df["Allocation"] = 0
        summary_df["MTM"] = 0

    numeric_cols = [
        "NFO Realized", "NFO Settlement", "BFO Realized", "BFO Settlement",
        "Total Realized", "Total Settlement", "Grand Total",
        "Allocation", "MTM",
    ]
    summary_df[numeric_cols] = (
        summary_df[numeric_cols].apply(pd.to_numeric, errors="coerce").fillna(0)
    )

    return summary_df.sort_values("UserID").reset_index(drop=True)


# ═════════════════════════════════════════════════════════════════════
# Portfolio exit analysis
# ═════════════════════════════════════════════════════════════════════
def process_portfolio_data(gridlog_file, summary_file):
    """
    Analyse portfolio exit reasons from GridLog + Summary files.

    Returns (final_df, output_filename).
    """
    if gridlog_file.name.endswith(".csv"):
        df_grid = pd.read_csv(gridlog_file)
    elif gridlog_file.name.endswith(".xlsx"):
        df_grid = pd.read_excel(gridlog_file)
    else:
        raise ValueError("Unsupported GridLog file type. Use CSV or Excel.")

    df_grid.columns = df_grid.columns.str.strip()

    mask = df_grid["Message"].str.contains(
        r"Combined SL:|Combined trail target:", case=False, na=False
    )
    filtered_grid = (
        df_grid.loc[mask, ["Message", "Option Portfolio", "Timestamp"]]
        .dropna(subset=["Option Portfolio"])
    )

    filtered_grid["MessageType"] = filtered_grid["Message"].str.extract(
        r"(Combined SL|Combined trail target)", flags=re.IGNORECASE
    )
    duplicate_mask = filtered_grid.duplicated(
        subset=["Option Portfolio", "MessageType"], keep=False
    )
    filtered_grid = filtered_grid[duplicate_mask]

    summary_grid = (
        filtered_grid.groupby("Option Portfolio")
        .agg({"Message": lambda x: ", ".join(x.unique()), "Timestamp": "max"})
        .reset_index()
        .rename(columns={"Message": "Reason", "Timestamp": "Time"})
    )

    xl = pd.ExcelFile(summary_file)
    summary_list = []

    for sheet_name in xl.sheet_names:
        if "legs" in sheet_name.lower():
            df_leg = xl.parse(sheet_name)
            df_leg.columns = df_leg.columns.str.strip()

            if {"Exit Type", "Portfolio Name", "Exit Time"}.issubset(df_leg.columns):
                onsqoff_df = df_leg[
                    df_leg["Exit Type"].astype(str).str.strip() == "OnSqOffTime"
                ]
                if not onsqoff_df.empty:
                    grouped = (
                        onsqoff_df.groupby("Portfolio Name")["Exit Time"]
                        .max()
                        .reset_index()
                    )
                    for _, row in grouped.iterrows():
                        summary_list.append({
                            "Option Portfolio": row["Portfolio Name"],
                            "Reason": "OnSqOffTime",
                            "Time": row["Exit Time"],
                        })

    summary_summary = pd.DataFrame(summary_list)
    final_df = pd.concat([summary_grid, summary_summary], ignore_index=True)
    final_df = (
        final_df.groupby("Option Portfolio")
        .agg({"Reason": lambda x: ", ".join(sorted(set(x))), "Time": "last"})
        .reset_index()
    )

    # ── Find "AllLegsCompleted" portfolios ───────────────────────────
    completed_list = []
    grid_portfolios = df_grid["Option Portfolio"].dropna().unique()

    for sheet_name in xl.sheet_names:
        if "legs" in sheet_name.lower():
            df_leg = xl.parse(sheet_name)
            df_leg.columns = df_leg.columns.str.strip()

            if "Portfolio Name" in df_leg.columns and "Status" in df_leg.columns:
                for portfolio, group in df_leg.groupby("Portfolio Name"):
                    if (
                        portfolio not in final_df["Option Portfolio"].values
                        and portfolio in grid_portfolios
                    ):
                        statuses = group["Status"].astype(str).str.strip().unique()
                        if len(statuses) == 1 and statuses[0].lower() == "completed":
                            reason_text = "AllLegsCompleted"
                            exit_time_to_use = None
                            if "Exit Time" in group.columns:
                                for exit_time, exit_type in zip(
                                    group["Exit Time"],
                                    group.get("Exit Type", []),
                                ):
                                    if pd.isna(exit_time):
                                        continue
                                    normalized = (
                                        str(exit_time).replace(".", ":").strip()
                                    )
                                    matching_rows = df_grid[
                                        (df_grid["Option Portfolio"] == portfolio)
                                        & (
                                            df_grid["Timestamp"]
                                            .astype(str)
                                            .str.contains(normalized)
                                        )
                                    ]
                                    if not matching_rows.empty:
                                        reason_text += f", {str(exit_type).strip()}"
                                        exit_time_to_use = exit_time
                                        break
                            completed_list.append({
                                "Option Portfolio": portfolio,
                                "Reason": reason_text,
                                "Time": exit_time_to_use,
                            })

    if completed_list:
        completed_df = pd.DataFrame(completed_list)
        final_df = pd.concat([final_df, completed_df], ignore_index=True)

    # ── Clean reasons ────────────────────────────────────────────────
    def _clean_reason(text):
        if pd.isna(text):
            return text
        text = str(text)
        match = re.search(
            r"(Combined SL: [^ ]+ hit|Combined Trail Target: [^ ]+ hit)",
            text,
            re.IGNORECASE,
        )
        if match:
            return match.group(1)
        if "AllLegsCompleted" in text:
            text = (
                text.replace("AllLegsCompleted,", "")
                .replace("AllLegsCompleted", "")
                .strip()
            )
        return text.strip()

    final_df["Reason"] = final_df["Reason"].apply(_clean_reason)

    # ── Generate output filename ─────────────────────────────────────
    filename = gridlog_file.name
    match = re.search(r"(\d{1,2}\s+[A-Za-z]{3}\s+\d{4})", filename)
    if match:
        raw_date = match.group(1)
        parts = raw_date.split()
        formatted_date = f"{parts[0]} {parts[1].lower()}"
    else:
        formatted_date = "unknown_date"
    output_filename = f"completed portfolio of {formatted_date}.csv"

    final_df["Time"] = final_df["Time"].astype(str).str.strip().replace("nan", None)
    return final_df, output_filename
